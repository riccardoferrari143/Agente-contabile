#!/usr/bin/env python3
"""
Agente Contabile Autonomo - Versione Completa
pip install pdfplumber openpyxl pandas fpdf2
"""
import os, re, json, glob, datetime
import pdfplumber
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from fpdf import FPDF

CATEGORIE_COSTI = [
    "Affitto e locazioni","Utenze (luce, gas, acqua, tel)",
    "Personale e collaboratori","Fornitori e acquisti",
    "Marketing e pubblicita","Trasporti e trasferte",
    "Spese bancarie e commissioni","Assicurazioni",
    "Materiali e attrezzature","Manutenzione e riparazioni",
    "Consulenze e servizi professionali","Imposte e tasse",
    "Ammortamenti","Spese generali e diverse",
]
CATEGORIE_RICAVI = [
    "Vendita prodotti","Prestazioni di servizi",
    "Canoni e abbonamenti","Interessi attivi","Altri ricavi",
]
PIANO_CONTI = {
    "110":"Cassa","120":"Banca c/c","130":"Crediti vs clienti",
    "140":"Crediti IVA","210":"Immobilizzazioni materiali",
    "220":"Fondo ammortamento","310":"Debiti vs fornitori",
    "320":"IVA a debito","330":"Debiti tributari",
    "410":"Capitale sociale","510":"Vendita prodotti",
    "520":"Prestazioni di servizi","530":"Canoni e abbonamenti",
    "540":"Altri ricavi","610":"Affitto e locazioni",
    "620":"Utenze","630":"Personale e collaboratori",
    "640":"Fornitori e acquisti","650":"Marketing",
    "660":"Trasporti e trasferte","670":"Spese bancarie",
    "680":"Assicurazioni","690":"Materiali e attrezzature",
    "700":"Manutenzione","710":"Consulenze professionali",
    "720":"Imposte e tasse","730":"Ammortamenti","740":"Spese generali",
}
CATEGORIA_CONTO = {
    "Affitto e locazioni":"610","Utenze (luce, gas, acqua, tel)":"620",
    "Personale e collaboratori":"630","Fornitori e acquisti":"640",
    "Marketing e pubblicita":"650","Trasporti e trasferte":"660",
    "Spese bancarie e commissioni":"670","Assicurazioni":"680",
    "Materiali e attrezzature":"690","Manutenzione e riparazioni":"700",
    "Consulenze e servizi professionali":"710","Imposte e tasse":"720",
    "Ammortamenti":"730","Spese generali e diverse":"740",
    "Vendita prodotti":"510","Prestazioni di servizi":"520",
    "Canoni e abbonamenti":"530","Interessi attivi":"540","Altri ricavi":"540",
}
KEYWORDS = {
    "Affitto e locazioni":["affitto","locazione","canone locazione","rent"],
    "Utenze (luce, gas, acqua, tel)":["enel","eni gas","acqua","telecom","tim","vodafone","wind","fastweb","utenza","bolletta","luce","gas"],
    "Personale e collaboratori":["stipendio","salario","collaboratore","dipendente","busta paga","compenso","inps","contributi"],
    "Fornitori e acquisti":["fornitore","acquisto","merce","materiale","ordine","fattura acquisto"],
    "Marketing e pubblicita":["marketing","pubblicita","advertising","google ads","facebook ads","sponsorizzazione","promozione"],
    "Trasporti e trasferte":["carburante","autostrada","pedaggio","taxi","treno","aereo","hotel","trasferta","trasporto","benzina"],
    "Spese bancarie e commissioni":["commissione","spesa bancaria","bonifico","pos","carta di credito","interessi passivi","banco"],
    "Assicurazioni":["assicurazione","polizza","premio assicurativo","infortuni"],
    "Materiali e attrezzature":["attrezzatura","computer","software","hardware","ufficio","cancelleria","stampante"],
    "Manutenzione e riparazioni":["manutenzione","riparazione","assistenza tecnica","pulizie"],
    "Consulenze e servizi professionali":["consulenza","commercialista","avvocato","notaio","studio","professionale"],
    "Imposte e tasse":["iva","irpef","ires","irap","imposta","tassa","tributo","agenzia entrate","f24"],
    "Ammortamenti":["ammortamento","quota ammortamento"],
    "Vendita prodotti":["vendita","prodotto venduto","merce venduta"],
    "Prestazioni di servizi":["prestazione","servizio reso","fattura emessa","onorario"],
    "Canoni e abbonamenti":["canone","abbonamento","subscription","rinnovo"],
    "Interessi attivi":["interessi attivi","proventi finanziari"],
    "Altri ricavi":["rimborso","accredito","provento","entrata"],
}

def classifica(descrizione, importo):
    desc = descrizione.lower()
    for cat, kws in KEYWORDS.items():
        for kw in kws:
            if kw in desc: return cat
    return "Altri ricavi" if importo >= 0 else "Spese generali e diverse"

def parse_importo(raw):
    raw = str(raw).strip().replace(" ","")
    if "," in raw and "." in raw:
        if raw.rindex(",") > raw.rindex("."): raw = raw.replace(".","").replace(",",".")
        else: raw = raw.replace(",","")
    elif "," in raw: raw = raw.replace(",",".")
    try: return float(raw)
    except: return 0.0

def parse_data(s):
    for fmt in ("%d/%m/%Y","%Y-%m-%d","%d-%m-%Y","%m/%d/%Y","%d/%m/%y"):
        try: return datetime.datetime.strptime(str(s).strip(),fmt).date()
        except: pass
    return None

def nuovo_movimento(data,desc,importo,aliq_iva=22.0,scadenza="",stato="",ndoc="",tipo="",fonte="",commessa=""):
    imp = importo
    imponibile = round(imp/(1+aliq_iva/100),2) if aliq_iva>0 else imp
    iva_imp = round(abs(imp)-abs(imponibile),2)
    if not stato: stato = "pagato" if imp>=0 else "da_pagare"
    if not tipo: tipo = "fattura_emessa" if imp>=0 else "fattura_ricevuta"
    return {"data":data,"descrizione":desc,"importo":imp,"imponibile":imponibile,
            "aliquota_iva":aliq_iva,"iva_importo":iva_imp,"scadenza":scadenza,
            "stato_pagamento":stato,"numero_documento":ndoc,"tipo_documento":tipo,
            "fonte":fonte,"categoria":classifica(desc,imp),"commessa":commessa}

def estrai_da_pdf(path):
    testo=""
    with pdfplumber.open(path) as pdf:
        for page in pdf.pages:
            t=page.extract_text()
            if t: testo+=t+"\n"
    importo=None
    for pat in [r"(?:totale|total|importo)\D{0,20}([\d\.]+[,\.][\d]{2})",
                r"([\d\.]+[,\.][\d]{2})\s*EUR",r"EUR\s*([\d\.]+[,\.][\d]{2})"]:
        m=re.search(pat,testo,re.IGNORECASE)
        if m: importo=-abs(parse_importo(m.group(1))); break
    aliq=22.0
    m2=re.search(r"IVA\s*(22|10|4)\s*%",testo,re.IGNORECASE)
    if m2: aliq=float(m2.group(1))
    data=None
    for pat in [r"(\d{2})[/\-\.](\d{2})[/\-\.](\d{4})",r"(\d{4})[/\-\.](\d{2})[/\-\.](\d{2})"]:
        m=re.search(pat,testo)
        if m:
            g=m.groups()
            try: data=f"{g[2]}/{g[1]}/{g[0]}" if len(g[0])==4 else f"{g[0]}/{g[1]}/{g[2]}"; break
            except: pass
    fornitore=os.path.splitext(os.path.basename(path))[0]
    for riga in testo.splitlines():
        riga=riga.strip()
        if len(riga)>3 and not riga.isdigit(): fornitore=riga[:60]; break
    imp=importo if importo else 0.0
    return nuovo_movimento(data or datetime.date.today().strftime("%d/%m/%Y"),
                           f"Fattura: {fornitore}",imp,aliq,fonte=os.path.basename(path))

def leggi_movimenti(path):
    righe=[]
    try:
        df=pd.read_csv(path,sep=None,engine="python",dtype=str) if path.endswith(".csv") else pd.read_excel(path,dtype=str)
    except Exception as e:
        print(f"  [ATTENZIONE] {e}"); return righe
    df.columns=[c.strip().lower() for c in df.columns]
    col_data=next((c for c in df.columns if "data" in c and "scad" not in c),None)
    col_desc=next((c for c in df.columns if any(k in c for k in ["descr","causale","nota","note","detail"])),None)
    col_imp=next((c for c in df.columns if any(k in c for k in ["importo","amount","valore"])),None)
    col_iva=next((c for c in df.columns if "iva" in c and "aliq" in c),None)
    col_scad=next((c for c in df.columns if "scad" in c),None)
    col_stato=next((c for c in df.columns if "stato" in c),None)
    col_ndoc=next((c for c in df.columns if "numero" in c or "n_doc" in c),None)
    col_tipo=next((c for c in df.columns if "tipo" in c),None)
    col_comm=next((c for c in df.columns if any(k in c for k in ["commessa","cantiere","progetto","job"])),None)
    if not col_imp: print(f"  [ATTENZIONE] Colonna importo non trovata"); return righe
    for _,row in df.iterrows():
        imp=parse_importo(str(row.get(col_imp,"0")))
        desc=str(row.get(col_desc,"")) if col_desc else ""
        data_str=str(row.get(col_data,"")) if col_data else datetime.date.today().strftime("%d/%m/%Y")
        aliq=parse_importo(str(row.get(col_iva,"22"))) if col_iva else 22.0
        scad=str(row.get(col_scad,"")) if col_scad else ""
        stato=str(row.get(col_stato,"")).lower() if col_stato else ""
        ndoc=str(row.get(col_ndoc,"")) if col_ndoc else ""
        tipo=str(row.get(col_tipo,"")).lower() if col_tipo else ""
        comm=str(row.get(col_comm,"")).strip() if col_comm else ""
        righe.append(nuovo_movimento(data_str,desc,imp,aliq,scad,stato,ndoc,tipo,os.path.basename(path),comm))
    return righe

def leggi_cespiti(path):
    if not os.path.exists(path): return []
    try:
        df=pd.read_excel(path,dtype=str)
        df.columns=[c.strip().lower() for c in df.columns]
        return [{"descrizione":str(r.get("descrizione","")),
                 "data_acquisto":str(r.get("data_acquisto","")),
                 "costo":parse_importo(str(r.get("costo","0"))),
                 "vita_utile":int(parse_importo(str(r.get("vita_utile_anni","5")))),
                 "aliquota_perc":parse_importo(str(r.get("aliquota_perc","20")))}
                for _,r in df.iterrows()]
    except Exception as e:
        print(f"  [ATTENZIONE] Errore cespiti: {e}"); return []

def calcola_ammortamenti(cespiti,anno):
    righe=[]
    for c in cespiti:
        d=parse_data(c["data_acquisto"])
        if not d: continue
        aliq=c["aliquota_perc"]/100 if c["aliquota_perc"]>1 else c["aliquota_perc"]
        quota=round(c["costo"]*aliq,2)
        anni=anno-d.year
        fondo=round(min(c["costo"],quota*anni),2)
        residuo=round(max(0,c["costo"]-fondo),2)
        righe.append({"descrizione":c["descrizione"],"data_acquisto":c["data_acquisto"],
                      "costo_storico":c["costo"],"aliquota_perc":c["aliquota_perc"],
                      "quota_annua":quota,"fondo_amm_prec":round(fondo-quota,2) if anni>0 else 0.0,
                      "quota_esercizio":quota if anni<c["vita_utile"] else 0.0,
                      "fondo_amm_fin":fondo,"valore_residuo":residuo})
    return righe

def inserisci_budget():
    print("\n"+"="*60+"\n  BUDGET MENSILE (INVIO per saltare)\n"+"="*60)
    budget={}
    for cat in CATEGORIE_COSTI:
        val=input(f"  {cat}: EUR ").strip()
        if val:
            try: budget[cat]=parse_importo(val)
            except: pass
    return budget

def carica_budget(path):
    if os.path.exists(path):
        with open(path,"r",encoding="utf-8") as f: return json.load(f)
    return {}

def salva_budget(budget,path):
    with open(path,"w",encoding="utf-8") as f: json.dump(budget,f,ensure_ascii=False,indent=2)

BLU="FF1F3864"; BLU_C="FFD9E1F2"; VERDE="FFE2EFDA"; ROSSO="FFFCE4D6"; GIALLO="FFFFF2CC"

def tb():
    t=Side(style="thin",color="BFBFBF")
    return Border(left=t,right=t,top=t,bottom=t)

def sh(cell):
    cell.font=Font(name="Calibri",bold=True,color="FFFFFF",size=11)
    cell.fill=PatternFill("solid",fgColor=BLU)
    cell.alignment=Alignment(horizontal="center",vertical="center",wrap_text=True)
    cell.border=tb()

def sc(cell,bg=None,bold=False):
    cell.font=Font(name="Calibri",size=10,bold=bold)
    cell.border=tb()
    cell.alignment=Alignment(vertical="center")
    if bg: cell.fill=PatternFill("solid",fgColor=bg)

def cw(ws,d):
    for col,w in d.items(): ws.column_dimensions[col].width=w

def foglio_movimenti(wb,movimenti):
    ws=wb.active; ws.title="Movimenti"
    hdrs=["Data","Descrizione","Importo EUR","Imponibile","Aliq. IVA %","IVA EUR",
          "Categoria","Commessa","Tipo Doc.","N. Documento","Scadenza","Stato Pag.","Fonte"]
    ws.append(hdrs)
    for cell in ws[1]: sh(cell)
    for m in movimenti:
        bg=VERDE if m["importo"]>=0 else None
        ws.append([m["data"],m["descrizione"],m["importo"],m["imponibile"],
                   m["aliquota_iva"],m["iva_importo"],m["categoria"],m.get("commessa",""),
                   m["tipo_documento"],m["numero_documento"],m["scadenza"],m["stato_pagamento"],m["fonte"]])
        for cell in ws[ws.max_row]: sc(cell,bg)
    cw(ws,{"A":12,"B":40,"C":13,"D":13,"E":11,"F":11,"G":32,"H":20,"I":18,"J":14,"K":12,"L":12,"M":22})

def foglio_prima_nota(wb,movimenti):
    ws=wb.create_sheet("Prima Nota")
    hdrs=["Data","Descrizione","Conto Dare","Nome Conto Dare","Conto Avere","Nome Conto Avere","Importo EUR"]
    ws.append(hdrs)
    for cell in ws[1]: sh(cell)
    for m in movimenti:
        imp=abs(m["importo"]); cat=m["categoria"]
        cec=CATEGORIA_CONTO.get(cat,"740"); nec=PIANO_CONTI.get(cec,cat)
        if m["importo"]>=0:
            dare=("130",PIANO_CONTI["130"]); avere=(cec,nec)
        else:
            dare=(cec,nec); avere=("310",PIANO_CONTI["310"])
        bg=VERDE if m["importo"]>=0 else None
        ws.append([m["data"],m["descrizione"],dare[0],dare[1],avere[0],avere[1],imp])
        for cell in ws[ws.max_row]: sc(cell,bg)
    cw(ws,{"A":12,"B":40,"C":10,"D":28,"E":10,"F":28,"G":13})

def foglio_registro_fatture(wb,movimenti):
    for titolo,lista in [("Fatture Emesse",[m for m in movimenti if m["importo"]>=0]),
                          ("Fatture Ricevute",[m for m in movimenti if m["importo"]<0])]:
        ws=wb.create_sheet(titolo)
        hdrs=["N. Doc.","Data","Controparte","Imponibile EUR","IVA %","IVA EUR","Totale EUR","Scadenza","Stato"]
        ws.append(hdrs)
        for cell in ws[1]: sh(cell)
        ti=tv=tt=0.0
        for i,m in enumerate(lista,1):
            ia=abs(m["imponibile"]); va=abs(m["iva_importo"]); ta=abs(m["importo"])
            stato=m["stato_pagamento"]
            bg=VERDE if stato=="pagato" else (ROSSO if stato=="scaduto" else GIALLO)
            ws.append([m["numero_documento"] or f"{i:04d}",m["data"],m["descrizione"],
                       ia,m["aliquota_iva"],va,ta,m["scadenza"],stato.upper()])
            for cell in ws[ws.max_row]: sc(cell,bg)
            ti+=ia; tv+=va; tt+=ta
        ws.append(["","TOTALE","",ti,"",tv,tt,"",""])
        for cell in ws[ws.max_row]: sc(cell,BLU_C,bold=True)
        cw(ws,{"A":10,"B":12,"C":35,"D":15,"E":8,"F":12,"G":13,"H":12,"I":12})

def foglio_iva(wb,movimenti):
    ws=wb.create_sheet("Gestione IVA")
    ws.append(["LIQUIDAZIONE IVA"])
    ws["A1"].font=Font(name="Calibri",bold=True,size=13,color="FFFFFF")
    ws["A1"].fill=PatternFill("solid",fgColor=BLU)
    ws.merge_cells("A1:D1"); ws.row_dimensions[1].height=22
    ws.append([])
    hdrs=["Voce","Imponibile EUR","Aliquota IVA","IVA EUR"]
    ws.append(hdrs)
    for cell in ws[3]: sh(cell)
    id_=sum(abs(m["iva_importo"]) for m in movimenti if m["importo"]>=0)
    ic_=sum(abs(m["iva_importo"]) for m in movimenti if m["importo"]<0)
    saldo=round(id_-ic_,2)
    ir=sum(abs(m["imponibile"]) for m in movimenti if m["importo"]>=0)
    iac=sum(abs(m["imponibile"]) for m in movimenti if m["importo"]<0)
    for voce,imp,iva in [("IVA a debito (su vendite)",ir,id_),("IVA a credito (su acquisti)",iac,ic_)]:
        ws.append([voce,imp,"22%",iva])
        for cell in ws[ws.max_row]: sc(cell)
    ws.append([])
    ws.append(["SALDO IVA (da versare)","","",saldo if saldo>0 else 0.0])
    ws.append(["CREDITO IVA (da recuperare)","","",abs(saldo) if saldo<0 else 0.0])
    for r in [ws.max_row-1,ws.max_row]:
        bg=ROSSO if saldo>0 else VERDE
        for cell in ws[r]: sc(cell,bg,bold=True)
    ws.append([]); ws.append(["DETTAGLIO PER ALIQUOTA"])
    ws.append(["Aliquota","N. Operazioni","Imponibile","IVA"])
    for cell in ws[ws.max_row]: sh(cell)
    df=pd.DataFrame(movimenti)
    for aliq,grp in df.groupby("aliquota_iva"):
        ws.append([f"{aliq}%",len(grp),round(grp["imponibile"].abs().sum(),2),round(grp["iva_importo"].abs().sum(),2)])
        for cell in ws[ws.max_row]: sc(cell)
    cw(ws,{"A":35,"B":16,"C":14,"D":14})

def foglio_scadenzario(wb,movimenti):
    ws=wb.create_sheet("Scadenzario")
    oggi=datetime.date.today()
    hdrs=["Data Doc.","Descrizione","Importo EUR","Scadenza","Giorni Mancanti","Stato","Tipo"]
    ws.append(hdrs)
    for cell in ws[1]: sh(cell)
    aperti=[m for m in movimenti if m.get("stato_pagamento","").lower() in ("da_pagare","da pagare","aperto","")]
    if not aperti: ws.append(["Nessuna scadenza aperta."]); return
    for m in sorted(aperti,key=lambda x: parse_data(x.get("scadenza","")) or datetime.date.max):
        scad=parse_data(m.get("scadenza",""))
        gg=(scad-oggi).days if scad else ""
        if isinstance(gg,int) and gg<0: stato="SCADUTO"; bg=ROSSO
        elif isinstance(gg,int) and gg<=7: stato="IN SCADENZA"; bg=GIALLO
        else: stato="APERTO"; bg=None
        tipo="DA INCASSARE" if m["importo"]>0 else "DA PAGARE"
        ws.append([m["data"],m["descrizione"],abs(m["importo"]),m.get("scadenza",""),gg,stato,tipo])
        for cell in ws[ws.max_row]: sc(cell,bg)
    cw(ws,{"A":12,"B":40,"C":13,"D":12,"E":15,"F":14,"G":14})

def foglio_riepilogo(wb,movimenti,budget):
    ws=wb.create_sheet("Riepilogo Categorie")
    ws.append(["Categoria","Tipo","Totale EUR","Budget EUR","Scostamento EUR","Stato"])
    for cell in ws[1]: sh(cell)
    df=pd.DataFrame(movimenti)
    for _,row in df.groupby("categoria")["importo"].sum().reset_index().iterrows():
        cat=row["categoria"]; tot=round(row["importo"],2)
        tipo="RICAVO" if tot>=0 else "COSTO"
        bud=budget.get(cat)
        if bud and tipo=="COSTO":
            scost=round(abs(tot)-bud,2); stato="SFORATO" if scost>0 else "OK"
            bg=ROSSO if scost>0 else VERDE
        else: scost=""; stato=""; bg=None
        ws.append([cat,tipo,tot,bud or "",scost,stato])
        for cell in ws[ws.max_row]: sc(cell,bg)
    cw(ws,{"A":38,"B":10,"C":14,"D":14,"E":18,"F":12})

def foglio_conto_economico(wb,movimenti):
    ws=wb.create_sheet("Conto Economico")
    ws.append(["Voce","Importo EUR"])
    for cell in ws[1]: sh(cell)
    df=pd.DataFrame(movimenti)
    tr=sum(m["importo"] for m in movimenti if m["importo"]>=0)
    tc=sum(m["importo"] for m in movimenti if m["importo"]<0)
    utile=round(tr+tc,2)
    righe=[("RICAVI",None)]
    for cat,val in df[df["importo"]>=0].groupby("categoria")["importo"].sum().items():
        righe.append((f"  {cat}",round(val,2)))
    righe.append(("TOTALE RICAVI",round(tr,2)))
    righe+=[("",None),("COSTI",None)]
    for cat,val in df[df["importo"]<0].groupby("categoria")["importo"].sum().items():
        righe.append((f"  {cat}",round(val,2)))
    righe.append(("TOTALE COSTI",round(tc,2)))
    righe+=[("",None),("UTILE / PERDITA",utile)]
    for voce,val in righe:
        ws.append([voce,val])
        r=ws.max_row
        if voce in ("TOTALE RICAVI","TOTALE COSTI","UTILE / PERDITA","RICAVI","COSTI"):
            for cell in ws[r]:
                cell.font=Font(name="Calibri",bold=True,size=11)
                if voce=="UTILE / PERDITA":
                    cell.fill=PatternFill("solid",fgColor=VERDE if utile>=0 else ROSSO)
        else:
            for cell in ws[r]: sc(cell)
    cw(ws,{"A":40,"B":16})

def foglio_bilancio(wb,movimenti,cespiti_amm):
    ws=wb.create_sheet("Bilancio Patrimoniale")
    tr=sum(m["importo"] for m in movimenti if m["importo"]>=0)
    tc=sum(m["importo"] for m in movimenti if m["importo"]<0)
    utile=round(tr+tc,2)
    cred_cli=sum(abs(m["importo"]) for m in movimenti if m["importo"]>=0 and m.get("stato_pagamento","") in ("da_pagare",""))
    deb_for=sum(abs(m["importo"]) for m in movimenti if m["importo"]<0 and m.get("stato_pagamento","") in ("da_pagare",""))
    imm_nette=sum(c["valore_residuo"] for c in cespiti_amm)
    fondo_amm=sum(c["fondo_amm_fin"] for c in cespiti_amm)
    costo_imm=sum(c["costo_storico"] for c in cespiti_amm)
    id_=sum(abs(m["iva_importo"]) for m in movimenti if m["importo"]>=0)
    ic_=sum(abs(m["iva_importo"]) for m in movimenti if m["importo"]<0)
    saldo_iva=round(id_-ic_,2)
    ws.append(["ATTIVO","EUR","","PASSIVO","EUR"])
    for col,c in [(1,"A1"),(4,"D1"),(2,"B1"),(5,"E1")]:
        cell=ws[c]
        cell.font=Font(name="Calibri",bold=True,color="FFFFFF",size=12)
        cell.fill=PatternFill("solid",fgColor=BLU)
    attivo=[("ATTIVO FISSO",None),("  Immob. materiali (costo)",costo_imm),
            ("  (-) Fondo ammortamento",-fondo_amm),("  Immob. nette",imm_nette),
            ("",""),("ATTIVO CIRCOLANTE",None),("  Crediti vs clienti",cred_cli),
            ("  Credito IVA",max(0,-saldo_iva)),("  Liquidita",max(0,tr-abs(tc)))]
    passivo=[("PATRIMONIO NETTO",None),("  Utile / Perdita",utile),
             ("",""),("PASSIVO A BREVE",None),("  Debiti vs fornitori",deb_for),
             ("  IVA a debito",max(0,saldo_iva)),("",""),("",""),("","")]
    for (va,vav),(vp,vpv) in zip(attivo,passivo):
        ws.append([va,vav if vav is not None else "","",vp,vpv if vpv is not None else ""])
        r=ws.max_row
        ba=va.strip() in ("ATTIVO FISSO","ATTIVO CIRCOLANTE","  Immob. nette")
        bp=vp.strip() in ("PATRIMONIO NETTO","PASSIVO A BREVE")
        for col in [1,2]: sc(ws.cell(r,col),bold=ba)
        for col in [4,5]: sc(ws.cell(r,col),bold=bp)
    ta=round(imm_nette+cred_cli+max(0,-saldo_iva)+max(0,tr-abs(tc)),2)
    tp=round(utile+deb_for+max(0,saldo_iva),2)
    ws.append(["TOTALE ATTIVO",ta,"","TOTALE PASSIVO + PN",tp])
    for cell in ws[ws.max_row]: sc(cell,BLU_C,bold=True)
    cw(ws,{"A":38,"B":14,"C":4,"D":38,"E":14})

def foglio_ammortamenti(wb,cespiti_amm):
    ws=wb.create_sheet("Ammortamenti")
    hdrs=["Descrizione","Data Acquisto","Costo Storico EUR","Aliquota %",
          "Quota Annua EUR","Fondo Amm. Prec.","Quota Esercizio","Fondo Amm. Finale","Valore Residuo EUR"]
    ws.append(hdrs)
    for cell in ws[1]: sh(cell)
    tq=tf=tr_=0.0
    for c in cespiti_amm:
        ws.append([c["descrizione"],c["data_acquisto"],c["costo_storico"],c["aliquota_perc"],
                   c["quota_annua"],c["fondo_amm_prec"],c["quota_esercizio"],c["fondo_amm_fin"],c["valore_residuo"]])
        for cell in ws[ws.max_row]: sc(cell)
        tq+=c["quota_esercizio"]; tf+=c["fondo_amm_fin"]; tr_+=c["valore_residuo"]
    ws.append(["TOTALE","","","",tq,"",tq,tf,tr_])
    for cell in ws[ws.max_row]: sc(cell,BLU_C,bold=True)
    cw(ws,{"A":35,"B":14,"C":18,"D":11,"E":15,"F":17,"G":17,"H":18,"I":18})

class PDF(FPDF):
    def header(self):
        self.set_font("Helvetica","B",14)
        self.set_fill_color(31,56,100); self.set_text_color(255,255,255)
        self.cell(0,12,"  REPORT CONTABILE COMPLETO",new_x="LMARGIN",new_y="NEXT",fill=True)
        self.set_text_color(0,0,0); self.set_font("Helvetica","",9)
        self.cell(0,6,f"  Generato il {datetime.date.today().strftime('%d/%m/%Y')}",new_x="LMARGIN",new_y="NEXT")
        self.ln(4)
    def footer(self):
        self.set_y(-15); self.set_font("Helvetica","I",8)
        self.set_text_color(128,128,128); self.cell(0,10,f"Pagina {self.page_no()}",align="C")
    def titolo(self,t):
        self.set_font("Helvetica","B",12); self.set_fill_color(217,225,242)
        self.set_text_color(0,0,0); self.cell(0,8,t,new_x="LMARGIN",new_y="NEXT",fill=True); self.ln(2)

def crea_report_pdf(movimenti,budget,cespiti_amm,path_output):
    pdf=PDF(); pdf.set_auto_page_break(auto=True,margin=15); pdf.add_page()
    df=pd.DataFrame(movimenti)
    tr=sum(m["importo"] for m in movimenti if m["importo"]>=0)
    tc=sum(m["importo"] for m in movimenti if m["importo"]<0)
    utile=tr+tc
    id_=sum(abs(m["iva_importo"]) for m in movimenti if m["importo"]>=0)
    ic_=sum(abs(m["iva_importo"]) for m in movimenti if m["importo"]<0)
    saldo_iva=id_-ic_

    pdf.titolo("CONTO ECONOMICO")
    for label,val,color in [("Totale Ricavi",tr,(46,117,182)),("Totale Costi",abs(tc),(192,0,0))]:
        pdf.set_text_color(*color); pdf.set_font("Helvetica","B",10)
        pdf.cell(110,7,f"  {label}:"); pdf.cell(0,7,f"EUR {val:,.2f}",new_x="LMARGIN",new_y="NEXT")
    pdf.set_text_color(0,0,0); pdf.set_font("Helvetica","B",11)
    pdf.set_fill_color(226,239,218) if utile>=0 else pdf.set_fill_color(252,228,214)
    pdf.cell(110,8,"  Utile / Perdita:"); pdf.cell(0,8,f"EUR {utile:,.2f}",fill=True,new_x="LMARGIN",new_y="NEXT")
    pdf.ln(4)

    pdf.titolo("LIQUIDAZIONE IVA")
    pdf.set_font("Helvetica","",10); pdf.set_text_color(0,0,0)
    pdf.cell(110,7,"  IVA a debito (vendite):"); pdf.cell(0,7,f"EUR {id_:,.2f}",new_x="LMARGIN",new_y="NEXT")
    pdf.cell(110,7,"  IVA a credito (acquisti):"); pdf.cell(0,7,f"EUR {ic_:,.2f}",new_x="LMARGIN",new_y="NEXT")
    pdf.set_font("Helvetica","B",11)
    label_iva="Saldo IVA (da versare)" if saldo_iva>0 else "Credito IVA (da recuperare)"
    pdf.cell(110,8,f"  {label_iva}:"); pdf.cell(0,8,f"EUR {abs(saldo_iva):,.2f}",new_x="LMARGIN",new_y="NEXT")
    pdf.ln(4)

    pdf.titolo("RIEPILOGO PER CATEGORIA")
    pdf.set_font("Helvetica","B",9); pdf.set_fill_color(31,56,100); pdf.set_text_color(255,255,255)
    for txt,w,al in [("Categoria",80,"L"),("Tipo",20,"C"),("Totale EUR",30,"R"),("Budget EUR",25,"R"),("Stato",25,"C")]:
        kw={"new_x":"LMARGIN","new_y":"NEXT"} if txt=="Stato" else {}
        pdf.cell(w,7,txt,border=1,fill=True,align=al,**kw)
    pdf.set_text_color(0,0,0)
    for _,row in df.groupby("categoria")["importo"].sum().reset_index().iterrows():
        cat=row["categoria"]; tot=round(row["importo"],2)
        tipo="RICAVO" if tot>=0 else "COSTO"; bud=budget.get(cat)
        if bud and tipo=="COSTO" and abs(tot)>bud: stato="SFORATO"; pdf.set_fill_color(252,228,214)
        elif bud and tipo=="COSTO": stato="OK"; pdf.set_fill_color(226,239,218)
        else: stato="-"; pdf.set_fill_color(255,255,255)
        pdf.set_font("Helvetica","",8)
        pdf.cell(80,6,f"  {cat}",border=1,fill=True)
        pdf.cell(20,6,tipo,border=1,fill=True,align="C")
        pdf.cell(30,6,f"{tot:,.2f}",border=1,fill=True,align="R")
        pdf.cell(25,6,f"{bud:,.2f}" if bud else "-",border=1,fill=True,align="R")
        pdf.cell(25,6,stato,border=1,fill=True,align="C",new_x="LMARGIN",new_y="NEXT")
    pdf.ln(4)

    aperti=[m for m in movimenti if m.get("stato_pagamento","").lower() in ("da_pagare","") and m.get("scadenza","")]
    if aperti:
        pdf.add_page(); pdf.titolo("SCADENZARIO PAGAMENTI APERTI")
        oggi=datetime.date.today()
        pdf.set_font("Helvetica","B",9); pdf.set_fill_color(31,56,100); pdf.set_text_color(255,255,255)
        for txt,w in [("Descrizione",65),("EUR",25),("Scadenza",25),("Giorni",20),("Stato",35)]:
            kw={"new_x":"LMARGIN","new_y":"NEXT"} if txt=="Stato" else {}
            pdf.cell(w,7,txt,border=1,fill=True,**kw)
        pdf.set_text_color(0,0,0)
        for m in sorted(aperti,key=lambda x: parse_data(x.get("scadenza","")) or datetime.date.max):
            scad=parse_data(m.get("scadenza",""))
            gg=(scad-oggi).days if scad else 0
            if gg<0: stato="SCADUTO"; pdf.set_fill_color(252,228,214)
            elif gg<=7: stato="IN SCADENZA"; pdf.set_fill_color(255,242,204)
            else: stato="APERTO"; pdf.set_fill_color(255,255,255)
            pdf.set_font("Helvetica","",8)
            pdf.cell(65,6,f"  {m['descrizione'][:38]}",border=1,fill=True)
            pdf.cell(25,6,f"{abs(m['importo']):,.2f}",border=1,fill=True,align="R")
            pdf.cell(25,6,m.get("scadenza",""),border=1,fill=True,align="C")
            pdf.cell(20,6,str(gg),border=1,fill=True,align="C")
            pdf.cell(35,6,stato,border=1,fill=True,align="C",new_x="LMARGIN",new_y="NEXT")

    if cespiti_amm:
        pdf.add_page(); pdf.titolo("REGISTRO AMMORTAMENTI")
        pdf.set_font("Helvetica","B",9); pdf.set_fill_color(31,56,100); pdf.set_text_color(255,255,255)
        for txt,w in [("Cespite",65),("Costo EUR",25),("Aliq. %",20),("Quota EUR",25),("Residuo EUR",30)]:
            kw={"new_x":"LMARGIN","new_y":"NEXT"} if txt=="Residuo EUR" else {}
            pdf.cell(w,7,txt,border=1,fill=True,**kw)
        pdf.set_text_color(0,0,0)
        for c in cespiti_amm:
            pdf.set_fill_color(255,255,255); pdf.set_font("Helvetica","",8)
            pdf.cell(65,6,f"  {c['descrizione'][:38]}",border=1,fill=True)
            pdf.cell(25,6,f"{c['costo_storico']:,.2f}",border=1,fill=True,align="R")
            pdf.cell(20,6,f"{c['aliquota_perc']}%",border=1,fill=True,align="C")
            pdf.cell(25,6,f"{c['quota_esercizio']:,.2f}",border=1,fill=True,align="R")
            pdf.cell(30,6,f"{c['valore_residuo']:,.2f}",border=1,fill=True,align="R",new_x="LMARGIN",new_y="NEXT")

    pdf.output(path_output)

def foglio_commesse(wb,movimenti):
    ws=wb.create_sheet("Riepilogo Commesse")
    ws.append(["RIEPILOGO PER COMMESSA / CANTIERE"])
    ws["A1"].font=Font(name="Calibri",bold=True,size=13,color="FFFFFF")
    ws["A1"].fill=PatternFill("solid",fgColor=BLU)
    ws.merge_cells("A1:G1"); ws.row_dimensions[1].height=22
    ws.append([])
    hdrs=["Commessa","Ricavi EUR","Costi EUR","Margine EUR","Margine %","N. Movimenti","Stato"]
    ws.append(hdrs)
    for cell in ws[3]: sh(cell)
    df=pd.DataFrame(movimenti)
    comm_list=sorted(df["commessa"].unique())
    for comm in comm_list:
        nome=comm if comm else "(senza commessa)"
        grp=df[df["commessa"]==comm]
        ric=round(grp[grp["importo"]>=0]["importo"].sum(),2)
        cos=round(grp[grp["importo"]<0]["importo"].sum(),2)
        mar=round(ric+cos,2)
        pct=round(mar/ric*100,1) if ric>0 else 0.0
        bg=VERDE if mar>=0 else ROSSO
        stato="IN UTILE" if mar>=0 else "IN PERDITA"
        ws.append([nome,ric,abs(cos),mar,f"{pct}%",len(grp),stato])
        for cell in ws[ws.max_row]: sc(cell,bg)
    # Totale
    tr=round(df[df["importo"]>=0]["importo"].sum(),2)
    tc=round(df[df["importo"]<0]["importo"].sum(),2)
    tot=round(tr+tc,2)
    ws.append(["TOTALE",tr,abs(tc),tot,f"{round(tot/tr*100,1) if tr>0 else 0}%",len(df),""])
    for cell in ws[ws.max_row]: sc(cell,BLU_C,bold=True)
    cw(ws,{"A":30,"B":14,"C":14,"D":14,"E":12,"F":14,"G":12})


SEZIONI = [
    ("1",  "Tutti i report (completo)",           "tutti"),
    ("2",  "Solo Ricavi",                          "ricavi"),
    ("3",  "Solo Costi",                           "costi"),
    ("4",  "Conto Economico",                      "conto_economico"),
    ("5",  "IVA (liquidazione)",                   "iva"),
    ("6",  "Scadenzario pagamenti",                "scadenzario"),
    ("7",  "Registro Fatture",                     "registro"),
    ("8",  "Prima Nota (partita doppia)",           "prima_nota"),
    ("9",  "Bilancio Patrimoniale",                "bilancio"),
    ("10", "Ammortamenti Cespiti",                 "ammortamenti"),
    ("11", "Movimenti dettaglio",                  "movimenti"),
    ("12", "Riepilogo per Categoria + Budget",     "riepilogo"),
    ("13", "Riepilogo per Commessa / Cantiere",     "commesse"),
]

def menu_sezioni():
    print("\n"+"="*60)
    print("  COSA VUOI GENERARE?")
    print("="*60)
    for num,label,_ in SEZIONI:
        print(f"  [{num:>2}] {label}")
    print("="*60)
    scelta=input("  Inserisci il numero (es. 4 per Conto Economico): ").strip()
    for num,label,key in SEZIONI:
        if scelta==num: return key,label
    return "tutti","Tutti i report (completo)"

def main():
    print("\n"+"="*60+"\n       AGENTE CONTABILE AUTONOMO - VERSIONE COMPLETA\n"+"="*60)
    radice=input("\nCartella radice del progetto (INVIO = cartella corrente): ").strip()
    if not radice: radice=os.getcwd()
    radice=os.path.expanduser(radice)
    cf=os.path.join(radice,"fatture"); cd=os.path.join(radice,"dati")
    co=os.path.join(radice,"output"); pc=os.path.join(radice,"cespiti.xlsx")
    os.makedirs(co,exist_ok=True)
    pb=os.path.join(radice,"budget.json")
    budget=carica_budget(pb)
    if budget:
        print(f"\nBudget caricato.")
        if input("Vuoi modificarlo? (s/N): ").strip().lower()=="s":
            budget=inserisci_budget(); salva_budget(budget,pb)
    else:
        print("\nNessun budget trovato.")
        if input("Vuoi impostare un budget mensile? (S/n): ").strip().lower()!="n":
            budget=inserisci_budget(); salva_budget(budget,pb)
    print("\n"+"="*60+"\n  FILTRO PERIODO (lascia vuoto per tutto)\n"+"="*60)
    dds=input("  Da data (es. 01/04/2026): ").strip()
    das=input("  A data  (es. 30/06/2026): ").strip()
    dd=parse_data(dds); da=parse_data(das)
    if dd or da: print(f"  Filtro: {dds or '...'} - {das or '...'}")
    else: print("  Nessun filtro - elaboro tutto.")
    # Filtro commessa
    tutte_commesse=[]
    print("\n"+"="*60+"\n  FILTRO COMMESSA (lascia vuoto per tutte)\n"+"="*60)
    print("  (potrai filtrare dopo aver caricato i dati - premi INVIO per ora)")
    commessa_filtro=input("  Commessa/Cantiere: ").strip().lower()

    def inp(ds):
        if not dd and not da: return True
        d=parse_data(str(ds))
        if not d: return True
        if dd and d<dd: return False
        if da and d>da: return False
        return True
    movimenti=[]
    os.makedirs(cf,exist_ok=True)
    pdfs=glob.glob(os.path.join(cf,"*.pdf"))
    if pdfs:
        print(f"\nTrovo {len(pdfs)} fatture PDF...")
        for f in pdfs:
            print(f"  Leggo: {os.path.basename(f)}")
            try:
                d=estrai_da_pdf(f)
                if inp(d["data"]): movimenti.append(d); print(f"    -> {d['categoria']} | EUR {d['importo']:.2f}")
                else: print("    -> Saltato (fuori periodo)")
            except Exception as e: print(f"    [ERRORE] {e}")
    else: print(f"\nNessuna fattura PDF in '{cf}'")
    os.makedirs(cd,exist_ok=True)
    dfs=glob.glob(os.path.join(cd,"*.xlsx"))+glob.glob(os.path.join(cd,"*.xls"))+glob.glob(os.path.join(cd,"*.csv"))
    if dfs:
        print(f"\nTrovo {len(dfs)} file dati...")
        for f in dfs:
            print(f"  Leggo: {os.path.basename(f)}")
            r=leggi_movimenti(f); filt=[x for x in r if inp(x["data"])]
            movimenti.extend(filt); print(f"    -> {len(filt)} movimenti (su {len(r)} totali)")
    else: print(f"\nNessun file Excel/CSV in '{cd}'")
    # Mostra commesse trovate e applica filtro
    commesse_trovate=sorted(set(m.get("commessa","") for m in movimenti if m.get("commessa","")))
    if commesse_trovate:
        print(f"\n  Commesse trovate: {', '.join(commesse_trovate)}")
        if not commessa_filtro:
            commessa_filtro=input("  Filtra per commessa (INVIO = tutte): ").strip().lower()
        if commessa_filtro:
            movimenti=[m for m in movimenti if m.get("commessa","").lower()==commessa_filtro]
            print(f"  Filtro applicato: '{commessa_filtro}' -> {len(movimenti)} movimenti")
    if not movimenti:
        print("\n[ATTENZIONE] Nessun movimento trovato."); input("\nPremi INVIO per uscire..."); return
    anno=(da or datetime.date.today()).year
    craw=leggi_cespiti(pc); camm=calcola_ammortamenti(craw,anno)
    if camm:
        qk=sum(c["quota_esercizio"] for c in camm)
        movimenti.append(nuovo_movimento(f"31/12/{anno}",f"Ammortamenti {anno}",-qk,0.0,stato="pagato",tipo="rettifica",fonte="cespiti.xlsx"))
        print(f"\nAmmortamenti: {len(camm)} cespiti, quota EUR {qk:,.2f}")
    tr=sum(m["importo"] for m in movimenti if m["importo"]>=0)
    tc=sum(m["importo"] for m in movimenti if m["importo"]<0)
    utile=tr+tc
    id_=sum(abs(m["iva_importo"]) for m in movimenti if m["importo"]>=0)
    ic_=sum(abs(m["iva_importo"]) for m in movimenti if m["importo"]<0)
    siva=id_-ic_
    print("\n"+"="*60+"\n  RIEPILOGO\n"+"="*60)
    print(f"  Movimenti      : {len(movimenti)}")
    print(f"  Totale ricavi  : EUR {tr:,.2f}")
    print(f"  Totale costi   : EUR {abs(tc):,.2f}")
    print(f"  Utile/Perdita  : EUR {utile:,.2f}")
    print(f"  IVA a debito   : EUR {id_:,.2f}")
    print(f"  IVA a credito  : EUR {ic_:,.2f}")
    print(f"  Saldo IVA      : EUR {siva:,.2f} {'(da versare)' if siva>0 else '(credito)'}")

    sezione,label_sezione=menu_sezioni()
    print(f"\n  Genero: {label_sezione}")

    ts=datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
    slug=sezione.replace(" ","_")
    pe=os.path.join(co,f"report_{slug}_{ts}.xlsx")
    pp=os.path.join(co,f"report_{slug}_{ts}.pdf")

    # Filtra movimenti per sezione se richiesto
    mov_ricavi=[m for m in movimenti if m["importo"]>=0]
    mov_costi=[m for m in movimenti if m["importo"]<0]

    print(f"\nGenerazione report Excel...")
    wb=openpyxl.Workbook()

    if sezione in ("tutti","movimenti"):
        foglio_movimenti(wb,movimenti)
    if sezione in ("tutti","prima_nota"):
        foglio_prima_nota(wb,movimenti)
    if sezione in ("tutti","registro"):
        foglio_registro_fatture(wb,movimenti)
    if sezione in ("tutti","iva"):
        foglio_iva(wb,movimenti)
    if sezione in ("tutti","scadenzario"):
        foglio_scadenzario(wb,movimenti)
    if sezione in ("tutti","riepilogo"):
        foglio_riepilogo(wb,movimenti,budget)
    if sezione in ("tutti","commesse"):
        foglio_commesse(wb,movimenti)
    if sezione in ("tutti","conto_economico"):
        foglio_conto_economico(wb,movimenti)
    if sezione in ("tutti","bilancio"):
        foglio_bilancio(wb,movimenti,camm)
    if sezione in ("tutti","ammortamenti") and camm:
        foglio_ammortamenti(wb,camm)
    if sezione=="ammortamenti" and not camm:
        print("  [ATTENZIONE] Nessun cespite trovato in cespiti.xlsx")

    # Fogli solo ricavi o solo costi
    if sezione=="ricavi":
        foglio_movimenti(wb,mov_ricavi)
        foglio_conto_economico(wb,mov_ricavi)
    if sezione=="costi":
        foglio_movimenti(wb,mov_costi)
        foglio_conto_economico(wb,mov_costi)

    # Rimuovi il foglio vuoto iniziale se ce ne sono altri
    if len(wb.sheetnames)>1 and wb.sheetnames[0]=="Sheet":
        del wb["Sheet"]

    wb.save(pe); print(f"  Salvato: {pe}")

    print(f"Generazione report PDF...")
    if sezione in ("tutti","conto_economico","ricavi","costi","riepilogo","iva","scadenzario","ammortamenti"):
        crea_report_pdf(movimenti,budget,camm,pp); print(f"  Salvato: {pp}")
    else:
        print(f"  (Report PDF non disponibile per questa sezione - usa il file Excel)")

    print("\n"+"="*60+"\n  ELABORAZIONE COMPLETATA\n  Output in: "+co+"\n"+"="*60)
    input("\nPremi INVIO per uscire...")

if __name__=="__main__":
    main()
