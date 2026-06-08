"""
Agente Contabile SaaS - App principale Streamlit
"""
import sys, os
sys.path.insert(0, os.path.dirname(__file__))

import streamlit as st
from core.database import init_db
from core.auth import init_session, login, registra, logout, require_login
from core.database import (get_aziende, crea_azienda, get_azienda, elimina_azienda,
                            salva_movimenti, get_movimenti, cancella_movimenti,
                            salva_report, get_report)
from core.stripe_payments import PIANI, limiti_piano

# Import engine contabile
sys.path.insert(0, os.path.join(os.path.dirname(__file__), ".."))
from agente_contabile import (leggi_movimenti, estrai_da_pdf, calcola_ammortamenti,
                               leggi_cespiti, foglio_movimenti, foglio_prima_nota,
                               foglio_registro_fatture, foglio_iva, foglio_scadenzario,
                               foglio_riepilogo, foglio_conto_economico, foglio_bilancio,
                               foglio_ammortamenti, foglio_commesse, crea_report_pdf,
                               nuovo_movimento)

import openpyxl, pandas as pd, plotly.express as px
import datetime, tempfile, io

# ── CONFIG ──────────────────────────────────────────────
st.set_page_config(
    page_title="Agente Contabile",
    page_icon="📊",
    layout="wide",
    initial_sidebar_state="expanded"
)

init_db()
init_session()

# ── CSS PERSONALIZZATO ───────────────────────────────────
st.markdown("""
<style>
    .main-header {font-size:2rem; font-weight:700; color:#1F3864; margin-bottom:0}
    .sub-header  {font-size:1rem; color:#666; margin-bottom:1.5rem}
    .card        {background:#f8f9fa; border-radius:12px; padding:1.2rem; border:1px solid #e0e0e0; margin-bottom:1rem}
    .metric-big  {font-size:1.8rem; font-weight:700}
    .green       {color:#2e7d32}
    .red         {color:#c62828}
    .badge-free  {background:#e3f2fd; color:#1565c0; padding:2px 10px; border-radius:20px; font-size:.8rem}
    .badge-base  {background:#e8f5e9; color:#2e7d32; padding:2px 10px; border-radius:20px; font-size:.8rem}
    .badge-pro   {background:#fff3e0; color:#e65100; padding:2px 10px; border-radius:20px; font-size:.8rem}
</style>
""", unsafe_allow_html=True)

# ══════════════════════════════════════════════════════════
# PAGINE DI AUTENTICAZIONE
# ══════════════════════════════════════════════════════════
def pagina_login():
    col1, col2, col3 = st.columns([1,2,1])
    with col2:
        st.markdown('<p class="main-header">📊 Agente Contabile</p>', unsafe_allow_html=True)
        st.markdown('<p class="sub-header">Contabilità professionale per partite IVA e società</p>', unsafe_allow_html=True)

        tab_login, tab_reg = st.tabs(["Accedi", "Registrati"])

        with tab_login:
            with st.form("form_login"):
                email = st.text_input("Email")
                password = st.text_input("Password", type="password")
                submit = st.form_submit_button("Accedi", use_container_width=True, type="primary")
            if submit:
                ok, result = login(email, password)
                if ok:
                    st.session_state.logged_in = True
                    st.session_state.utente = result
                    st.rerun()
                else:
                    st.error(result)

        with tab_reg:
            with st.form("form_reg"):
                nome = st.text_input("Nome completo")
                email_r = st.text_input("Email")
                pwd1 = st.text_input("Password", type="password")
                pwd2 = st.text_input("Conferma password", type="password")
                submit_r = st.form_submit_button("Registrati", use_container_width=True, type="primary")
            if submit_r:
                if pwd1 != pwd2:
                    st.error("Le password non coincidono.")
                else:
                    ok, msg = registra(email_r, nome, pwd1)
                    if ok:
                        st.success("Registrazione completata! Accedi con le tue credenziali.")
                    else:
                        st.error(msg)

        st.divider()
        st.markdown("**Piani disponibili:**")
        cols = st.columns(3)
        for i,(piano,info) in enumerate(PIANI.items()):
            with cols[i]:
                prezzo = f"EUR {info['prezzo']//100}/mese" if info['prezzo']>0 else "Gratuito"
                st.markdown(f"""<div class='card' style='text-align:center'>
                    <b>{info['nome']}</b><br>
                    <span style='font-size:1.3rem;font-weight:700'>{prezzo}</span><br>
                    <small>Fino a {info['aziende']} aziend{'a' if info['aziende']==1 else 'e'}<br>
                    {info['movimenti_mese']} movimenti/mese</small>
                </div>""", unsafe_allow_html=True)

# ══════════════════════════════════════════════════════════
# SIDEBAR
# ══════════════════════════════════════════════════════════
def sidebar():
    utente = st.session_state.utente
    piano = utente.get("piano","free")
    badge_class = f"badge-{piano}"
    with st.sidebar:
        st.markdown(f"**👤 {utente['nome']}**")
        st.markdown(f'<span class="{badge_class}">Piano {piano.upper()}</span>', unsafe_allow_html=True)
        st.divider()

        # Selezione azienda
        aziende = get_aziende(utente["id"])
        limiti = limiti_piano(piano)

        if aziende:
            nomi = [a["nome"] for a in aziende]
            idx = 0
            if st.session_state.azienda_id:
                ids = [a["id"] for a in aziende]
                if st.session_state.azienda_id in ids:
                    idx = ids.index(st.session_state.azienda_id)
            scelta = st.selectbox("🏢 Azienda attiva", nomi, index=idx)
            az = next(a for a in aziende if a["nome"]==scelta)
            st.session_state.azienda_id = az["id"]
            st.session_state.azienda_nome = az["nome"]

        st.divider()
        st.markdown("**Navigazione**")
        for label, key in [("🏠 Dashboard","dashboard"),("📤 Carica dati","upload"),
                            ("📊 Genera report","report"),("📋 Movimenti","movimenti"),
                            ("🏢 Aziende","aziende"),("💳 Abbonamento","abbonamento")]:
            if st.button(label, use_container_width=True,
                         type="primary" if st.session_state.page==key else "secondary"):
                st.session_state.page = key; st.rerun()
        st.divider()
        if st.button("🚪 Esci", use_container_width=True): logout()

# ══════════════════════════════════════════════════════════
# DASHBOARD
# ══════════════════════════════════════════════════════════
def pagina_dashboard():
    utente = st.session_state.utente
    st.markdown('<p class="main-header">🏠 Dashboard</p>', unsafe_allow_html=True)

    if not st.session_state.azienda_id:
        st.info("Nessuna azienda. Creane una dalla sezione **Aziende**.")
        return

    movimenti = get_movimenti(st.session_state.azienda_id)
    if not movimenti:
        st.info("Nessun movimento. Vai su **Carica dati** per importare i tuoi dati.")
        return

    df = pd.DataFrame(movimenti)
    tr = df[df["importo"]>=0]["importo"].sum()
    tc = df[df["importo"]<0]["importo"].sum()
    utile = tr + tc
    id_ = df[df["importo"]>=0]["iva_importo"].sum()
    ic_ = df[df["importo"]<0]["iva_importo"].sum()
    siva = id_ - ic_

    # KPI
    col1,col2,col3,col4 = st.columns(4)
    with col1:
        st.metric("💰 Ricavi totali", f"EUR {tr:,.2f}")
    with col2:
        st.metric("💸 Costi totali", f"EUR {abs(tc):,.2f}")
    with col3:
        delta_color = "normal" if utile >= 0 else "inverse"
        st.metric("📈 Utile/Perdita", f"EUR {utile:,.2f}", delta=f"{'▲' if utile>=0 else '▼'}")
    with col4:
        label_iva = "IVA da versare" if siva>0 else "Credito IVA"
        st.metric(f"🧾 {label_iva}", f"EUR {abs(siva):,.2f}")

    st.divider()
    col_left, col_right = st.columns(2)

    with col_left:
        st.subheader("Ricavi vs Costi per categoria")
        grp = df.groupby("categoria")["importo"].sum().reset_index()
        grp["tipo"] = grp["importo"].apply(lambda x: "Ricavo" if x>=0 else "Costo")
        grp["importo_abs"] = grp["importo"].abs()
        fig = px.bar(grp, x="importo_abs", y="categoria", color="tipo",
                     color_discrete_map={"Ricavo":"#2e7d32","Costo":"#c62828"},
                     orientation="h", labels={"importo_abs":"EUR","categoria":""})
        fig.update_layout(height=400, margin=dict(l=0,r=0,t=20,b=0))
        st.plotly_chart(fig, use_container_width=True)

    with col_right:
        st.subheader("Andamento mensile")
        df["mese"] = pd.to_datetime(df["data"], dayfirst=True, errors="coerce").dt.to_period("M").astype(str)
        mensile = df.groupby("mese").apply(
            lambda g: pd.Series({"Ricavi": g[g["importo"]>=0]["importo"].sum(),
                                  "Costi": g[g["importo"]<0]["importo"].abs().sum()})
        ).reset_index()
        fig2 = px.line(mensile, x="mese", y=["Ricavi","Costi"],
                       color_discrete_map={"Ricavi":"#2e7d32","Costi":"#c62828"},
                       labels={"value":"EUR","mese":"Mese","variable":""})
        fig2.update_layout(height=400, margin=dict(l=0,r=0,t=20,b=0))
        st.plotly_chart(fig2, use_container_width=True)

    # Scadenzario
    scaduti = [m for m in movimenti if m.get("stato_pagamento","") in ("da_pagare","") and m.get("scadenza","")]
    if scaduti:
        st.subheader(f"⏰ Scadenzario ({len(scaduti)} aperti)")
        df_sc = pd.DataFrame(scaduti)[["data","descrizione","importo","scadenza","stato_pagamento","commessa"]]
        st.dataframe(df_sc, use_container_width=True, hide_index=True)

# ══════════════════════════════════════════════════════════
# CARICA DATI
# ══════════════════════════════════════════════════════════
def pagina_upload():
    st.markdown('<p class="main-header">📤 Carica dati</p>', unsafe_allow_html=True)
    if not st.session_state.azienda_id:
        st.info("Seleziona o crea un'azienda prima."); return

    st.info(f"**Azienda:** {st.session_state.azienda_nome}")

    col1, col2 = st.columns(2)
    with col1:
        st.subheader("📊 Movimenti Excel / CSV")
        st.caption("Colonne accettate: data, descrizione, importo, aliquota_iva, scadenza, stato, commessa")
        file_excel = st.file_uploader("Carica file Excel o CSV", type=["xlsx","xls","csv"],
                                       accept_multiple_files=True, key="up_excel")
    with col2:
        st.subheader("📄 Fatture PDF")
        st.caption("L'agente estrae automaticamente importo, data e fornitore")
        file_pdf = st.file_uploader("Carica fatture PDF", type=["pdf"],
                                     accept_multiple_files=True, key="up_pdf")

    if file_excel or file_pdf:
        nuovi = []
        with st.spinner("Elaborazione in corso..."):
            for f in (file_excel or []):
                with tempfile.NamedTemporaryFile(suffix=os.path.splitext(f.name)[1], delete=False) as tmp:
                    tmp.write(f.read()); tmp_path = tmp.name
                movs = leggi_movimenti(tmp_path)
                os.unlink(tmp_path)
                nuovi.extend(movs)
                st.success(f"✅ {f.name}: {len(movs)} movimenti")
            for f in (file_pdf or []):
                with tempfile.NamedTemporaryFile(suffix=".pdf", delete=False) as tmp:
                    tmp.write(f.read()); tmp_path = tmp.name
                try:
                    m = estrai_da_pdf(tmp_path)
                    nuovi.append(m)
                    st.success(f"✅ {f.name}: EUR {m['importo']:,.2f}")
                except Exception as e:
                    st.error(f"❌ {f.name}: {e}")
                os.unlink(tmp_path)

        if nuovi:
            st.divider()
            st.write(f"**{len(nuovi)} movimenti pronti da importare**")
            df_prev = pd.DataFrame(nuovi)[["data","descrizione","importo","categoria","commessa"]]
            st.dataframe(df_prev, use_container_width=True, hide_index=True)

            col_a, col_b = st.columns(2)
            with col_a:
                if st.button("✅ Importa tutti", type="primary", use_container_width=True):
                    salva_movimenti(st.session_state.azienda_id, nuovi)
                    st.success(f"Importati {len(nuovi)} movimenti!")
                    st.rerun()
            with col_b:
                if st.button("🗑️ Cancella tutti i movimenti esistenti e reimporta", use_container_width=True):
                    cancella_movimenti(st.session_state.azienda_id)
                    salva_movimenti(st.session_state.azienda_id, nuovi)
                    st.success("Dati sostituiti.")
                    st.rerun()

# ══════════════════════════════════════════════════════════
# GENERA REPORT
# ══════════════════════════════════════════════════════════
def pagina_report():
    st.markdown('<p class="main-header">📊 Genera Report</p>', unsafe_allow_html=True)
    if not st.session_state.azienda_id:
        st.info("Seleziona un'azienda prima."); return

    utente = st.session_state.utente
    col1, col2, col3 = st.columns(3)
    with col1:
        da_str = st.text_input("Da data (es. 01/01/2026)", "")
    with col2:
        a_str  = st.text_input("A data  (es. 31/12/2026)", "")
    with col3:
        movimenti = get_movimenti(st.session_state.azienda_id)
        commesse = sorted(set(m.get("commessa","") for m in movimenti if m.get("commessa","")))
        commessa_scelta = st.selectbox("Commessa (opzionale)", ["Tutte"] + commesse)

    sezioni_opzioni = {
        "Tutti i report":"tutti","Solo Ricavi":"ricavi","Solo Costi":"costi",
        "Conto Economico":"conto_economico","IVA":"iva","Scadenzario":"scadenzario",
        "Registro Fatture":"registro","Prima Nota":"prima_nota",
        "Bilancio Patrimoniale":"bilancio","Ammortamenti":"ammortamenti",
        "Riepilogo Categorie":"riepilogo","Riepilogo Commesse":"commesse",
    }
    sezione_label = st.selectbox("Tipo di report", list(sezioni_opzioni.keys()))
    sezione = sezioni_opzioni[sezione_label]

    if st.button("🚀 Genera Report", type="primary", use_container_width=True):
        movimenti_filtrati = get_movimenti(
            st.session_state.azienda_id,
            da=da_str or None, a=a_str or None,
            commessa=commessa_scelta if commessa_scelta!="Tutte" else None
        )
        if not movimenti_filtrati:
            st.warning("Nessun movimento trovato con i filtri selezionati."); return

        with st.spinner("Generazione in corso..."):
            wb = openpyxl.Workbook()
            mov_r = [m for m in movimenti_filtrati if m["importo"]>=0]
            mov_c = [m for m in movimenti_filtrati if m["importo"]<0]

            if sezione in ("tutti","movimenti"): foglio_movimenti(wb, movimenti_filtrati)
            if sezione in ("tutti","prima_nota"): foglio_prima_nota(wb, movimenti_filtrati)
            if sezione in ("tutti","registro"): foglio_registro_fatture(wb, movimenti_filtrati)
            if sezione in ("tutti","iva"): foglio_iva(wb, movimenti_filtrati)
            if sezione in ("tutti","scadenzario"): foglio_scadenzario(wb, movimenti_filtrati)
            if sezione in ("tutti","riepilogo"): foglio_riepilogo(wb, movimenti_filtrati, {})
            if sezione in ("tutti","conto_economico"): foglio_conto_economico(wb, movimenti_filtrati)
            if sezione in ("tutti","bilancio"): foglio_bilancio(wb, movimenti_filtrati, [])
            if sezione in ("tutti","commesse"): foglio_commesse(wb, movimenti_filtrati)
            if sezione == "ricavi": foglio_movimenti(wb, mov_r); foglio_conto_economico(wb, mov_r)
            if sezione == "costi": foglio_movimenti(wb, mov_c); foglio_conto_economico(wb, mov_c)

            if len(wb.sheetnames)>1 and "Sheet" in wb.sheetnames: del wb["Sheet"]

            buf_xlsx = io.BytesIO()
            wb.save(buf_xlsx); buf_xlsx.seek(0)

            buf_pdf = None
            if sezione in ("tutti","conto_economico","ricavi","costi","riepilogo","iva","scadenzario","ammortamenti"):
                with tempfile.NamedTemporaryFile(suffix=".pdf", delete=False) as tmp:
                    crea_report_pdf(movimenti_filtrati, {}, [], tmp.name)
                    tmp_pdf_path = tmp.name
                with open(tmp_pdf_path, "rb") as f: buf_pdf = io.BytesIO(f.read())
                os.unlink(tmp_pdf_path)

        ts = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
        slug = sezione.replace(" ","_")
        nome_az = st.session_state.azienda_nome.replace(" ","_")[:20]
        fname_xlsx = f"report_{nome_az}_{slug}_{ts}.xlsx"
        fname_pdf  = f"report_{nome_az}_{slug}_{ts}.pdf"

        st.success("✅ Report generato!")
        col_dl1, col_dl2 = st.columns(2)
        with col_dl1:
            st.download_button("⬇️ Scarica Excel", buf_xlsx, fname_xlsx,
                               "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                               use_container_width=True)
        with col_dl2:
            if buf_pdf:
                buf_pdf.seek(0)
                st.download_button("⬇️ Scarica PDF", buf_pdf, fname_pdf,
                                   "application/pdf", use_container_width=True)

# ══════════════════════════════════════════════════════════
# MOVIMENTI
# ══════════════════════════════════════════════════════════
def pagina_movimenti():
    st.markdown('<p class="main-header">📋 Movimenti</p>', unsafe_allow_html=True)
    if not st.session_state.azienda_id:
        st.info("Seleziona un'azienda prima."); return

    movimenti = get_movimenti(st.session_state.azienda_id)
    if not movimenti:
        st.info("Nessun movimento. Vai su **Carica dati**."); return

    df = pd.DataFrame(movimenti)
    col1,col2,col3 = st.columns(3)
    with col1:
        filtro_tipo = st.selectbox("Tipo", ["Tutti","Ricavi","Costi"])
    with col2:
        commesse = ["Tutte"] + sorted(set(df["commessa"].dropna().unique()))
        filtro_comm = st.selectbox("Commessa", commesse)
    with col3:
        filtro_cat = st.selectbox("Categoria", ["Tutte"] + sorted(df["categoria"].dropna().unique().tolist()))

    df_vis = df.copy()
    if filtro_tipo == "Ricavi": df_vis = df_vis[df_vis["importo"]>=0]
    if filtro_tipo == "Costi":  df_vis = df_vis[df_vis["importo"]<0]
    if filtro_comm != "Tutte":  df_vis = df_vis[df_vis["commessa"]==filtro_comm]
    if filtro_cat  != "Tutte":  df_vis = df_vis[df_vis["categoria"]==filtro_cat]

    st.caption(f"{len(df_vis)} movimenti | Totale: EUR {df_vis['importo'].sum():,.2f}")
    cols_show = ["data","descrizione","importo","categoria","commessa","scadenza","stato_pagamento"]
    st.dataframe(df_vis[cols_show], use_container_width=True, hide_index=True)

# ══════════════════════════════════════════════════════════
# AZIENDE
# ══════════════════════════════════════════════════════════
def pagina_aziende():
    st.markdown('<p class="main-header">🏢 Le tue Aziende</p>', unsafe_allow_html=True)
    utente = st.session_state.utente
    piano = utente.get("piano","free")
    limiti = limiti_piano(piano)
    aziende = get_aziende(utente["id"])

    for az in aziende:
        with st.expander(f"🏢 {az['nome']} — P.IVA {az['partita_iva'] or 'N/D'}"):
            movs = get_movimenti(az["id"])
            st.caption(f"{len(movs)} movimenti | Tipo: {az['tipo']}")
            if st.button(f"Elimina {az['nome']}", key=f"del_{az['id']}"):
                elimina_azienda(az["id"], utente["id"])
                st.rerun()

    st.divider()
    if len(aziende) >= limiti["aziende"]:
        st.warning(f"Hai raggiunto il limite di {limiti['aziende']} aziend{'a' if limiti['aziende']==1 else 'e'} per il piano {piano.upper()}. Passa a un piano superiore.")
        return

    with st.expander("➕ Aggiungi nuova azienda"):
        with st.form("form_az"):
            nome = st.text_input("Nome azienda / Ragione sociale")
            piva = st.text_input("Partita IVA (opzionale)")
            tipo = st.selectbox("Tipo", ["ditta_individuale","srl","srl_semplificata","snc","sas","spa","professionista"])
            ok = st.form_submit_button("Crea azienda", type="primary")
        if ok and nome:
            aid = crea_azienda(utente["id"], nome, piva, tipo)
            st.session_state.azienda_id = aid
            st.session_state.azienda_nome = nome
            st.success(f"✅ Azienda '{nome}' creata!")
            st.rerun()

# ══════════════════════════════════════════════════════════
# ABBONAMENTO
# ══════════════════════════════════════════════════════════
def pagina_abbonamento():
    st.markdown('<p class="main-header">💳 Abbonamento</p>', unsafe_allow_html=True)
    utente = st.session_state.utente
    piano_attuale = utente.get("piano","free")

    st.info(f"Piano attuale: **{piano_attuale.upper()}**")
    st.divider()

    cols = st.columns(3)
    for i,(piano,info) in enumerate(PIANI.items()):
        with cols[i]:
            prezzo_str = f"EUR {info['prezzo']//100}/mese" if info['prezzo']>0 else "Gratuito"
            attuale = "✅ Piano attuale" if piano==piano_attuale else ""
            st.markdown(f"""<div class='card' style='text-align:center'>
                <b style='font-size:1.2rem'>{info['nome']}</b><br>
                <span style='font-size:1.5rem;font-weight:700'>{prezzo_str}</span><br><br>
                ✔ Fino a {info['aziende']} aziend{'a' if info['aziende']==1 else 'e'}<br>
                ✔ {info['movimenti_mese']} movimenti/mese<br>
                ✔ Tutti i report<br><br>
                <b>{attuale}</b>
            </div>""", unsafe_allow_html=True)
            if piano != piano_attuale and piano != "free":
                if st.button(f"Passa a {info['nome']}", key=f"btn_{piano}", use_container_width=True, type="primary"):
                    st.info("🔧 Configura STRIPE_SECRET_KEY e STRIPE_PRICE_IDS nel file .env per attivare i pagamenti.")

# ══════════════════════════════════════════════════════════
# ROUTER PRINCIPALE
# ══════════════════════════════════════════════════════════
if not st.session_state.logged_in:
    pagina_login()
else:
    sidebar()
    page = st.session_state.get("page","dashboard")
    if page == "dashboard":    pagina_dashboard()
    elif page == "upload":     pagina_upload()
    elif page == "report":     pagina_report()
    elif page == "movimenti":  pagina_movimenti()
    elif page == "aziende":    pagina_aziende()
    elif page == "abbonamento":pagina_abbonamento()
